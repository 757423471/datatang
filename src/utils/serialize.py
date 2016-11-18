# -*- coding: utf-8 -*-

import csv
import openpyxl as px

def load_csv(filename):
	content = []
	with open(filename, 'rb') as csvfile:
		csvreader = csv.reader(csvfile)
		for row in csvreader:
			content.append(row)
	return content

def dump_csv(content, filename):
	with open(filename, 'wb') as csvfile:
		csvwriter = csv.writer(csvfile)
		for row in content:
			csvwriter.writerow(row)

# TODO: filters workbook with conditions - 
# only return content if its sheet_title, column and row were valid
def load_xlsx(filename, sheet_titles=[], columns=[], rows=[]):
	wb = px.load_workbook(filename=filename)
	workbook = {}
	for ws in wb:
		if not ws.title in sheet_titles:
			continue
		workbook[ws.title] = []
		for row in ws.rows:
			vals = []
			for cell in row:
				vals.append(cell.value)
			workbook[ws.title].append(vals)
	return workbook


def dump_xlsx(workbook, filename, sheet_titles=[], columns=[], rows=[]):
	wb = px.Workbook()
	actived = False
	for sheet_title, sheet_content in workbook.items():
		if not sheet_title in sheet_titles:
			continue

		if actived:
			ws = create_sheet(title=sheet_title)
		else:
			ws = wb.active
			ws.title = sheet_title
			actived = True

		for row in sheet_content:
			ws.append(row)
	wb.save(filename=filename)
